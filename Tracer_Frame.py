import os
import os.path
import pickle
import shutil
import tkinter as tk
from datetime import datetime, timedelta
from itertools import takewhile, dropwhile
from tkinter import messagebox as msg
from tkinter import ttk, scrolledtext

import instaloader
import openpyxl
import pandas as pd
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
# If modifying these scopes, delete the file token.pickle.
from googleapiclient.http import MediaFileUpload
from instaloader import *


class TracerFrame(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        tk.Frame.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        parent.focus()

        self.L = instaloader.Instaloader(save_metadata=False, download_comments=False,
                                    download_video_thumbnails=False)

        self.usernames = self.get_usernames_to_work()
        self.usernames.sort()

        self.create_information_label()
        self.create_selector_for_usernames()
        self.create_download_button()
        self.create_status_label()
        self.create_scrolledtext_to_showout_processes()
        self.add_progress_bar_to_track_the_processes()
        self.create_button_to_open_excel_app()
        self.add_button_to_add_a_new_competitors()
        self.add_button_to_delete_competitors()
        self.entry_box_for_new_competitors()
        self.entry_box_for_deleting_competitors()

    def create_information_label(self):
        information_label = ttk.Label(self, text='Enter ig username:', font=('bold', 10))
        information_label.grid(row=0, column=0)

    def create_selector_for_usernames(self):
        self.username_selected = tk.StringVar()
        self.selector_combobox = ttk.Combobox(self, width=17, textvariable=self.username_selected)
        self.selector_combobox['values'] = self.usernames
        self.selector_combobox.grid(row=0, column=1, sticky='NSWE')
        self.selector_combobox.current(0)

    def create_download_button(self):
        download_button = ttk.Button(self, text='Start Download', command=self.download_and_save_information)
        download_button.grid(row=0, column=2, sticky='NSWE')

    def create_status_label(self):
        self.status_label = ttk.Label(self, text='Non Iniziato')
        self.status_label.grid(row=1, column=0, columnspan=3)

    def create_button_to_open_excel_app(self):
        self.open_excel_button = ttk.Button(self, text='Open excel tracer file', command=self.open_excel_file)
        self.open_excel_button.grid(row=4, column=0, pady=10)

    def create_scrolledtext_to_showout_processes(self):
        scrol_w = 50
        scrol_h = 3
        font = ('bold', 9)
        self.scr = scrolledtext.ScrolledText(self, width=scrol_w, height=scrol_h, wrap=tk.WORD, font=font)
        self.scr.grid(column=0, columnspan=3, sticky="NSWE")

    def add_progress_bar_to_track_the_processes(self):
        self.progress_bar = ttk.Progressbar(self, orient='horizontal', length=350, mode='determinate')
        self.progress_bar.grid(column=0, row=3, columnspan=3, pady=10)
        self.progress_bar['value'] = 0

    def add_button_to_add_a_new_competitors(self):
        self.button_add_competitors = ttk.Button(self, text='Add competitors', command=self.add_competitors_to_file)
        self.button_add_competitors.grid(row=4, column=1, pady=10)

    def add_button_to_delete_competitors(self):
        self.button_to_delete_competitors = ttk.Button(self, text='Delete competitors', command=self.delete_competitor_in_the_file)
        self.button_to_delete_competitors.grid(row=4, column=2)

    def entry_box_for_new_competitors(self):
        self.new_competitor = tk.StringVar()
        self.entry_new_competitor = ttk.Entry(self, textvariable=self.new_competitor)
        self.entry_new_competitor.grid(row=5, column=1, pady=10, padx=20)

    def entry_box_for_deleting_competitors(self):
        self.competitor_to_delete = tk.StringVar()
        self.entry_deleted_competitor = ttk.Entry(self, textvariable=self.competitor_to_delete)
        self.entry_deleted_competitor.grid(row=5, column=2, pady=10, padx=20)

    def download_and_save_information(self):
        self.progress_bar['value'] = 0
        self.progress_bar['maximum'] = (7*len(self.usernames))+2
        text = ''
        operazioni_effettuate = False
        self.open_excel_button.configure(state='disabled')
        self.button_add_competitors.configure(state='disabled')
        self.button_to_delete_competitors.configure(state='disabled')
        self.instaloader_login()  # ok
        self.run_progressbar()
        self.update()
        for ig_username in self.usernames:
            text = f"Operazioni in corso per l'account {ig_username}..."
            self.insert_processes_in_scrolledtext(text)
            self.run_progressbar()
            self.first_check_username(ig_username) #ok
            self.run_progressbar()

            is_profile_recovered = self.profile_recovery(ig_username) #ok
            self.run_progressbar()
            if is_profile_recovered:
                self.number_of_stories = self.get_number_of_stories() # ok
                self.run_progressbar()
                self.download_stories(ig_username) # ok
                self.run_progressbar()
                self.save_information_excel(ig_username) # ok
                self.run_progressbar()
                self.save_information_on_the_cloud(ig_username) # ok
                self.run_progressbar()
                operazioni_effettuate = True
            else:
                self.progress_bar['value'] = 0
                msg.showerror("Errore recupero profilo", "Il profilo non è stato recuperato riprova")

        if operazioni_effettuate:
            self.run_progressbar()
            text = 'Operazioni di download ed upload effettuate'
            self.insert_processes_in_scrolledtext(text)
            self.status_label.config(text=text)
            self.open_excel_button.configure(state='normal')
            self.button_add_competitors.configure(state='normal')
            self.button_to_delete_competitors.configure(state='normal')
            self.update()

        else:
            msg.showerror("Errore operazione", "Mi dispiace ma le operazioni non sono state effettuate con successo")

    def download_stories(self, ig_username):
        stories_downloaded = 0
        for story in self.L.get_stories(userids=self.profile):
            # story is a Story object
            for item in story.get_items():
                # item is a StoryItem object
                self.L.download_storyitem(item, f'{ig_username}_stories')
                stories_downloaded += 1
                self.status_label.config(text=f'Downloaded ({stories_downloaded}/{self.number_of_stories})')
                self.update()

        self.status_label.config(text='Scaricamento stories effettuato')
        self.update()

        # move file beetween folders
        directory_list = []
        for x in os.listdir('.'):
            if os.path.isdir(x) and (x != 'build') and (x != '.idea') and (x != '__pycache__') and (x != 'lib'):
                directory_list.append(x)

        inside_path_directory = os.path.join(os.getcwd(), f'{ig_username}_stories')

        if stories_downloaded > 0:
            for x in os.listdir(inside_path_directory):
                if not (os.path.exists(f'C:\\Users\\nerva\\Documents\\Firless\\Firless_competitors\\{ig_username}\\{ig_username}_stories\\{x}')):
                    shutil.move(f'{inside_path_directory}\\{x}',
                                f'C:\\Users\\nerva\\Documents\\Firless\\Firless_competitors\\{ig_username}\\{ig_username}_stories\\{x}')

            shutil.rmtree(inside_path_directory)

    def save_information_excel(self, ig_username):
        duplicated_overwritten = False
        row_to_overwritten = 0
        actual_data = datetime.now()
        actual_data = actual_data.strftime("%m/%d/%Y")
        excel_directory_path = os.path.join('C:\\', 'Users', 'nerva', 'Documents', 'Firless', 'Firless_data', 'competitors_ig_trace.xlsx')
        workbook = openpyxl.load_workbook(excel_directory_path)
        data_frame = pd.read_excel(excel_directory_path)
        data_frame = data_frame[(data_frame['Account']==ig_username) & (data_frame['Data']==actual_data)]
        number_row = data_frame.shape
        number_row_df = number_row[0]
        if number_row_df == 1:
            row_to_overwritten = data_frame['ID'].iloc[0] + 1
            duplicated_overwritten = True

        sheet1 = workbook['Dati Account']

        actual_number_of_rows = sum(1 for x in list(sheet1.rows))
        data_profile = Profile.from_username(self.L.context, ig_username)
        profile_followers = data_profile.followers
        profile_post = sum(1 for x in data_profile.get_posts())
        todays_posts = self.get_todays_number_of_posts(ig_username)
        profile_stories = self.L.get_stories(userids=self.profile)
        number_of_stories = [story.itemcount for story in profile_stories]
        if len(number_of_stories) > 0:
            number_of_stories = number_of_stories[0]
        else:
            number_of_stories = 0

        actual_data = datetime.now()
        actual_data = actual_data.strftime("%m/%d/%Y")
        list_of_data_to_insert = [actual_number_of_rows, actual_data, ig_username, number_of_stories,
                                  profile_followers, profile_post, todays_posts]
        list_of_data_to_insert_duplicated = [(row_to_overwritten-1), actual_data, ig_username, number_of_stories,
                                             profile_followers, profile_post, todays_posts]
        if duplicated_overwritten:
            for i in range(1, 8):
                sheet1.cell(row=row_to_overwritten, column=i).value = list_of_data_to_insert_duplicated[i-1]
        else:
            for i in range(1, 8):
                sheet1.cell(row=(actual_number_of_rows + 1), column=i).value = list_of_data_to_insert[i-1]
        workbook.save(excel_directory_path)
        self.status_label.config(text='Aggiornamento excel completato')
        self.update()

    def save_information_on_the_cloud(self, ig_username):
        SCOPES = ['https://www.googleapis.com/auth/drive']

        """Shows basic usage of the Drive v3 API.
            Prints the names and ids of the first 10 files the user has access to.
            """
        ig_username = ig_username
        creds = None
        # The file token.pickle stores the user's access and refresh tokens, and is
        # created automatically when the authorization flow completes for the first
        # time.
        if os.path.exists('token.pickle'):
            with open(os.path.join('C:\\', 'Users', 'nerva', 'PycharmProjects', 'IG_Competitors_Tracer', 'token.pickle'), 'rb') as token:
                creds = pickle.load(token)
        # If there are no (valid) credentials available, let the user log in.
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    'C:/Users/nerva/Documents/Firless/Firless_data/google_drive_new.json', SCOPES)
                creds = flow.run_local_server(port=0)
            # Save the credentials for the next run
            with open(os.path.join('C:\\', 'Users', 'nerva', 'PycharmProjects', 'IG_Competitors_Tracer', 'token.pickle'), 'wb') as token:
                pickle.dump(creds, token)

        service = build('drive', 'v3', credentials=creds)

        # Call the Drive v3 API
        results = service.files().list(
            pageSize=1000, fields="nextPageToken, files(id, name)").execute()
        items = results.get('files', [])

        folder_id = ''
        file_id = ''
        children_folder = ''
        if not items:
            print('No files found.')
        else:
            print('Files:')
            for item in items:
                if item['name'] == f'competitors ig trace.xlsx':
                    file_id = item['id']
                if item['name'] == f'Fir*less Data':
                    folder_id = item['id']
                if item['name'] == f'{ig_username}:stories':
                    children_folder = item['id']
                print(u'{0} ({1})'.format(item['name'], item['id']))

        # Eliminating the excel tracer file
        if file_id != '':
            eliminato = service.files().delete(fileId=file_id).execute()
            print(eliminato)
            print('File eliminato')

        # Uploading the excel tracer file
        print(folder_id)
        file_metadata = {
            'name': f'competitors ig trace.xlsx',
            'parents': [folder_id]
        }
        media = MediaFileUpload(
            f'C:/Users/nerva/Documents/Firless/Firless_data/competitors_ig_trace.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True)
        file = service.files().create(body=file_metadata,
                                      media_body=media,
                                      fields='id').execute()

        print('File ID: %s' % file.get('id'))
        print('File caricato')

        # work with stories
        all_stories = []
        results = service.files().list(
            pageSize=1000, fields="nextPageToken, files(id, name)",
            q=f"'{children_folder}' in parents").execute()
        items = results.get('files', [])
        print('Files:')
        for item in items:
            print(u'{0} ({1})'.format(item['name'], item['id']))
            all_stories.append(item['name'])

        file_to_add_to_drive = []
        for x in os.listdir(
                f'C:/Users/nerva/Documents/Firless/Firless_competitors/{ig_username}/{ig_username}_stories'):
            if x not in all_stories:
                file_to_add_to_drive.append(x)

        print(file_to_add_to_drive)

        for file in file_to_add_to_drive:
            # getting if the file is mp4 or jpg
            mime_type = ''
            file_mime_type = file[-3:]
            if file_mime_type == 'jpg':
                mime_type = 'image/jpeg'
            else:
                mime_type = 'video/mp4'

            file_metadata = {
                'name': f'{file}',
                'parents': [children_folder]
            }
            media = MediaFileUpload(
                f'C:/Users/nerva/Documents/Firless/Firless_competitors/{ig_username}/{ig_username}_stories/{file}',
                mimetype=mime_type,
                resumable=True)
            file = service.files().create(body=file_metadata,
                                          media_body=media,
                                          fields='id').execute()

            print('File ID: %s' % file.get('id'))
            print('File caricato')
        self.status_label.config(text=f"Caricamento sul cloud dell'account {ig_username} effettuato")
        self.update()

    def instaloader_login(self):
        self.user = 'alessandronerva'
        self.password= 'Juventus1982?A'

        self.L.login(self.user, self.password)

        self.status_label.configure(text='Login Effettuato')
        self.update()

    def first_check_username(self, ig_username):
        if ig_username.strip() == "":
            msg.showerror("Errore Username", "Non hai inserito un username corretto")

    def profile_recovery(self, ig_username):
        try:
            self.profile = [Profile.from_username(self.L.context, ig_username).userid]
            self.profile_name = Profile.from_username(self.L.context, ig_username).username

            self.status_label.configure(text='Recupero Profilo effettuato')
            self.update()
            return True
        except ProfileNotExistsException:
            msg.showerror("Errore username", "Hai inserito un username che non esiste")
            return False

    def get_number_of_stories(self):
        number_of_stories = 0
        profile = self.profile
        stories = self.L.get_stories(userids=profile)
        for story in stories:
            number_of_stories = story.itemcount

        return number_of_stories

    def get_todays_number_of_posts(self, ig_username):
        posts = instaloader.Profile.from_username(self.L.context, ig_username).get_posts()

        since = datetime.now().date()
        until = datetime.now().date() - timedelta(days=1)
        number_of_posts = 0

        for post in takewhile(lambda p: p.date.date() > until, dropwhile(lambda p: p.date.date() > since, posts)):
            number_of_posts += 1
            print(post.date)
            print(post.profile)

        return number_of_posts

    def open_excel_file(self):
        os.system("start C:/Users/nerva/Documents/Firless/Firless_data/competitors_ig_trace.xlsx")

    @staticmethod
    def get_usernames_to_work():
        file = open('C:/Users/nerva/Documents/Firless/Firless_data/competitors_ig_username.txt', 'r')
        text = file.read()
        text_split = text.split(';')

        return text_split

    def insert_processes_in_scrolledtext(self, value):
        print(value)
        self.scr.insert(tk.END, value + '\n')
        self.scr.yview(tk.END)

    def run_progressbar(self):
        self.progress_bar['value'] += 1
        self.progress_bar.update()

    def add_competitors_to_file(self):
        self.usernames = self.get_usernames_to_work()
        new_competitor = self.entry_new_competitor.get()
        recover_profile = self.profile_recovery(new_competitor)
        if recover_profile:
            file = open('C:/Users/nerva/Documents/Firless/Firless_data/competitors_ig_username.txt', 'a')
            file.write(f';{self.entry_new_competitor.get()}')
            file.close()
            self.usernames.append(self.entry_new_competitor.get())
            self.usernames.sort()
            self.selector_combobox.config(values=self.usernames)
            self.selector_combobox.current(0)
            self.update()

    def delete_competitor_in_the_file(self):
        deleted_competitor = self.competitor_to_delete.get()
        if deleted_competitor in self.usernames:
            file = open('C:/Users/nerva/Documents/Firless/Firless_data/competitors_ig_username.txt', 'w')
            self.usernames.remove(deleted_competitor)
            inizio = True
            for username in self.usernames:
                if inizio:
                    file.write(f'{username}')
                    inizio = False
                else:
                    file.write(f';{username}')
            file.close()
            self.selector_combobox.config(values=self.usernames)
            self.selector_combobox.current(0)
            self.update()
        else:
            msg.showerror('Errore username', 'Hai inserito un username attualmente non presente riprova')